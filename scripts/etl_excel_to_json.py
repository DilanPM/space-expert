# -*- coding: utf-8 -*-
"""
ETL : Database.xlsx  ->  src/data/*.json  (Space Expert / DG DEFIS D1)

Lit les 15 feuilles, NETTOIE (erreurs de formules #NAME?/#VALUE!, colonnes
dupliquees, lignes fusionnees, coordonnees GPS), FUSIONNE monde/EU, et ecrit
des JSON propres. Lance:  python scripts/etl_excel_to_json.py
"""
import json, re, os, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Database.xlsx")
OUT  = os.path.join(ROOT, "src", "data")
os.makedirs(OUT, exist_ok=True)

# ─── Helpers ────────────────────────────────────────────────────────────────
ERR_RE = re.compile(r"^#(NAME|VALUE|REF|DIV|N/A|NULL|NUM)[!?]?", re.I)

def is_err(v):
    return isinstance(v, str) and (v.startswith("#") and ("?" in v or "!" in v) or ERR_RE.match(v))

def s(v):
    """Clean string -> normalised text or None (drops formula errors / empties)."""
    if v is None: return None
    if is_err(v): return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    t = str(v).replace("\xa0", " ").replace("​", "")
    t = re.sub(r"\[\d+\]", "", t)          # strip wiki footnote markers [52]
    t = re.sub(r"[ \t]*\n[ \t]*", " · ", t)  # flatten line breaks
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t if t else None

def num(v):
    """Parse a numeric value; returns float/int, or None. Keeps ranges as None (use rng)."""
    if v is None or is_err(v): return None
    if isinstance(v, (int, float)): return v
    t = str(v).strip().replace(",", "")
    m = re.match(r"^-?\d+(\.\d+)?$", t)
    if m:
        f = float(t)
        return int(f) if f.is_integer() else f
    return None

def rng(v):
    """Keep a raw value that may be a number or a textual range ('2-9', '0.5–2')."""
    n = num(v)
    if n is not None: return n
    return s(v)

def bullets(v):
    """Restructure a dash/semicolon/bullet-delimited field into a clean list."""
    t = s(v)
    if not t: return None
    # source uses ' — ', '•', '·', ';' or a leading '— ' as item separators
    parts = re.split(r"\s*[—•·]\s*|\s*;\s+", t)
    parts = [p.strip(" -–—:").strip() for p in parts if p and p.strip(" -–—:")]
    # drop a leading label that is just a connector word
    return parts or [t]

def slug(t):
    if not t: return None
    t = str(t).lower().replace("+", " plus ")
    t = re.sub(r"[^a-z0-9]+", "-", t)
    return t.strip("-") or None

def dedupe_ids(records):
    """Guarantee unique 'id' across a record list (append -2, -3 on collision)."""
    seen = {}
    for r in records:
        base = r.get("id") or "item"
        if base in seen:
            seen[base] += 1
            r["id"] = f"{base}-{seen[base]}"
        else:
            seen[base] = 1
    return records

def parse_coords(v):
    """'5.236°N 52.775°W' -> [lon, lat] (GeoJSON order for react-simple-maps)."""
    t = s(v)
    if not t: return None
    # normalise weird degree chars
    t = t.replace("�", "°")
    m = re.findall(r"(-?\d+(?:\.\d+)?)\s*°?\s*([NSEW])", t, re.I)
    if len(m) < 2: return None
    lat = lon = None
    for val, hemi in m:
        f = float(val); hemi = hemi.upper()
        if hemi in ("N", "S"): lat = f if hemi == "N" else -f
        if hemi in ("E", "W"): lon = f if hemi == "E" else -f
    if lat is None or lon is None: return None
    return [round(lon, 5), round(lat, 5)]

def header_map(ws, row=1):
    """First-occurrence header -> col index (0-based). Drops dup columns."""
    hm = {}
    for r in ws.iter_rows(min_row=row, max_row=row, values_only=True):
        for i, h in enumerate(r):
            key = s(h)
            if key and key not in hm:
                hm[key] = i
    return hm

def get(row, hm, *names):
    for n in names:
        if n in hm and hm[n] < len(row):
            return row[hm[n]]
    return None

# ─── European footprint (for isEU detection) ────────────────────────────────
EU_COUNTRIES = {
    "france","germany","italy","spain","united kingdom","uk","portugal","netherlands",
    "belgium","switzerland","austria","poland","czech republic","czechia","greece",
    "luxembourg","romania","denmark","finland","sweden","norway","ireland","slovenia",
    "slovakia","hungary","bulgaria","croatia","estonia","latvia","lithuania","malta",
    "cyprus","iceland","ukraine","europe","eu","european union","esa",
}
def is_eu_country(c):
    t = (s(c) or "").lower()
    return any(k in t for k in EU_COUNTRIES)

wb = openpyxl.load_workbook(XLSX, data_only=True)
report = {}

def save(name, data):
    p = os.path.join(OUT, name)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    report[name] = len(data) if isinstance(data, list) else len(data.get("items", data))
    print(f"  ✓ {name:28} {report[name]} records")

# ─── 1. LAUNCHERS (ALL + EU merge) ──────────────────────────────────────────
def build_launchers():
    launchers = {}

    # World catalogue
    ws = wb["Launchers - ALL"]; hm = header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(get(row, hm, "ID luncher name"))
        if not name: continue
        key = name.lower()
        launchers[key] = {
            "id": slug(name), "name": name,
            "family": s(get(row, hm, "Launcher family")),
            "category": s(get(row, hm, "Launcher category")),
            "status": s(get(row, hm, "Launcher status")),
            "manufacturer": s(get(row, hm, "Launcher manufacturer")),
            "manufacturerCountry": s(get(row, hm, "Manufacturer country")),
            "priceCatalogueMUSD": num(get(row, hm, "Price catalogue (M$)")),
            "priceCatalogueYear": num(get(row, hm, "Price catalogue year")),
            "priceLEOkUSDkg": num(get(row, hm, "Specific price to LEO (K$/kg)")),
            "priceSSOkUSDkg": num(get(row, hm, "Specific price to SSO (K$/kg)")),
            "priceGTOkUSDkg": num(get(row, hm, "Specific price to GTO (K$/kg)")),
            "capLEO": num(get(row, hm, "Capacity to LEO (kg)")),
            "capSSO": num(get(row, hm, "Capacity to SSO (kg)")),
            "capGTO": num(get(row, hm, "Capacity to GTO (kg)")),
            "leoAlt": num(get(row, hm, "LEO Altitude (km)")),
            "isEU": False, "source": "all",
        }

    # EU enrichment sheet
    ws = wb["Launchers - EU"]; hm = header_map(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(get(row, hm, "ID luncher name"))
        if not name: continue
        key = name.lower()
        rec = launchers.get(key) or {"id": slug(name), "name": name, "source": "eu", "isEU": False}
        eu_flag = s(get(row, hm, "Europe"))
        country = s(get(row, hm, "Manufacturer country")) or rec.get("manufacturerCountry")
        enrich = {
            "abbreviation": s(get(row, hm, "Abbreviation")),
            "trl": num(get(row, hm, "TRL")),
            "reusability": s(get(row, hm, "Reusability")),
            "maidenLaunch": s(get(row, hm, "Maiden launch date")),
            "spaceportPad": s(get(row, hm, "Spaceport & Launch pad")),
            "spaceportCountry": s(get(row, hm, "Spaceport Country")),
            "maxLaunchRate": rng(get(row, hm, "Max launch rate (yr)")),
            "rideshare": s(get(row, hm, "Rideshare")),
            "fundings": s(get(row, hm, "Fundings")),
            "needs": s(get(row, hm, "Needs")),
            "techInfo": s(get(row, hm, "Technical relevant info")),
            "additionalInfo": s(get(row, hm, "Additional info")),
            "capGEO": num(get(row, hm, "Capacity to GEO (kg)")),
            "capEscape": num(get(row, hm, "Capacity to Escape (kg)")),
            "propStage1": s(get(row, hm, "Propellant - 1st stage")),
            "propStage2": s(get(row, hm, "Propellant - 2nd stage")),
            "propStage3": s(get(row, hm, "Propellant - 3rd stage")),
            "booster": s(get(row, hm, "Booster")),
        }
        for k, v in enrich.items():
            if v is not None: rec[k] = v
        # fill any missing core fields from EU sheet
        for k, hk in [("family","Launcher family"),("category","Launcher category"),
                      ("status","Launcher status"),("manufacturer","Launcher manufacturer"),
                      ("manufacturerCountry","Manufacturer country"),
                      ("priceCatalogueMUSD","Price catalogue (M$)"),
                      ("capLEO","Capacity to LEO (kg)"),("capSSO","Capacity to SSO (kg)"),
                      ("capGTO","Capacity to GTO (kg)")]:
            if not rec.get(k):
                val = num(get(row, hm, hk)) if k.startswith(("price","cap")) else s(get(row, hm, hk))
                if val is not None: rec[k] = val
        rec["isEU"] = bool(eu_flag) or is_eu_country(country)
        launchers[key] = rec

    # Final isEU pass by country, drop empties
    out = []
    for rec in launchers.values():
        if not rec.get("name"): continue
        if not rec.get("isEU"):
            rec["isEU"] = is_eu_country(rec.get("manufacturerCountry"))
        rec.setdefault("confidence", "internalDB")
        out.append(rec)
    out.sort(key=lambda r: (not r["isEU"], r.get("name","")))
    return dedupe_ids(out)

# ─── 2. LAUNCHER CLASSIFICATION ─────────────────────────────────────────────
def build_class():
    ws = wb["Launchers Class"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows: return []
    headers = [s(c) for c in rows[0]]
    out = []
    for r in rows[1:]:
        cells = [s(c) for c in r]
        if not any(cells): continue
        obj = {}
        for i, h in enumerate(headers):
            if h and i < len(cells) and cells[i] is not None:
                obj[h] = cells[i]
        if obj: out.append(obj)
    return out

# ─── 3. CSG LAUNCHPADS (group merged rows + coords) ─────────────────────────
def build_launchpads():
    ws = wb["Launchpads"]; hm = header_map(ws)
    pads, cur = [], None
    for row in ws.iter_rows(min_row=2, values_only=True):
        complex_name = s(get(row, hm, "Complex"))
        uses   = s(get(row, hm, "Uses"))
        info   = s(get(row, hm, "Information"))
        notab  = s(get(row, hm, "Notable Launches"))
        coords = parse_coords(get(row, hm, "Coordinates"))
        if complex_name:
            cur = {
                "id": slug(complex_name), "name": complex_name,
                "status": s(get(row, hm, "Status")),
                "tenant": s(get(row, hm, "Tenant")),
                "uses": [u for u in [uses] if u],
                "info": [i for i in [info] if i],
                "notableLaunches": [n for n in [notab] if n],
                "coords": coords, "spaceportId": "guiana-space-centre-csg",
            }
            pads.append(cur)
        elif cur:
            if uses:  cur["uses"].append(uses)
            if info:  cur["info"].append(info)
            if notab: cur["notableLaunches"].append(notab)
            if coords and not cur["coords"]: cur["coords"] = coords
    return pads

# ─── 4. SPACEPORTS ──────────────────────────────────────────────────────────
def build_spaceports():
    ws = wb["Spaceports"]; hm = header_map(ws)
    cols = [
        ("name","Spaceport Name"),("location","Location (city)"),
        ("country","Country / Territory"),("launchType","Launch Type"),
        ("partners","Primary launch partner(s)"),("status","Status"),
        ("status2025","Status (2025)"),("compatibility","Launcher compatibility"),
        ("euInterest","EU strategic interest"),("euFunding","EU funding / potential funding"),
        ("history","Historique (last launches / key milestones)"),
        ("nextSteps","Next steps (licenses / planned first orbital flights)"),
        ("advantages","Advantages"),("disadvantages","Disadvantages"),
    ]
    BULLET = {"partners", "compatibility", "euInterest", "euFunding", "history", "nextSteps", "advantages", "disadvantages"}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(get(row, hm, "Spaceport Name"))
        if not name: continue
        obj = {"id": slug(name)}
        for key, hk in cols:
            obj[key] = bullets(get(row, hm, hk)) if key in BULLET else s(get(row, hm, hk))
        obj["confidence"] = "internalDB"
        out.append(obj)
    return out

# ─── 5. PROPULSION (3 sheets combined) ──────────────────────────────────────
def build_propulsion():
    out = []
    sheets = [("Propu Use - Launcher","launcher-use"),
              ("Propu tech - Launcher","launcher-tech"),
              ("Propulsion - Sats","satellite")]
    for sheet, scope in sheets:
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows: continue
        headers = [s(c) for c in rows[0]]
        for r in rows[1:]:
            cells = [s(c) for c in r]
            if not any(cells): continue
            obj = {"scope": scope, "fields": {}}
            for i, h in enumerate(headers):
                if h and i < len(cells) and cells[i] is not None:
                    obj["fields"][h] = cells[i]
            if obj["fields"]:
                first = next(iter(obj["fields"].values()))
                obj["id"] = slug(f"{scope}-{first[:40]}")
                obj["title"] = first
                out.append(obj)
    return out

# ─── 6. CRITICAL MATERIALS (Chemicals) ──────────────────────────────────────
def build_materials():
    ws = wb["Chemicals"]; hm = header_map(ws)
    cols = [
        ("category","Category"),("material","Chemical / Material"),
        ("spaceUse","Main Space Use"),
        ("launcherRole","Launchers (role & criticality)"),
        ("satelliteRole","Satellites (role & criticality)"),
        ("euRegulation","EU Regulation (REACH / others)"),
        ("pfas","PFAS relevance"),
        ("suppliers","Main EU Suppliers (examples)"),
        ("supplierRisk","Supplier Risk (EU autonomy)"),
    ]
    # the assessment col header is long/truncated -> grab by prefix
    assess_key = next((k for k in hm if k.lower().startswith("assessment")), None)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = s(get(row, hm, "Chemical / Material"))
        if not mat: continue
        obj = {"id": slug(mat)}
        for key, hk in cols:
            obj[key] = s(get(row, hm, hk))
        obj["assessment"] = s(row[hm[assess_key]]) if assess_key and hm[assess_key] < len(row) else None
        out.append(obj)
    return dedupe_ids(out)

# ─── 7. GROUND INFRA ────────────────────────────────────────────────────────
def build_ground_infra():
    ws = wb["Ground Infra"]; hm = header_map(ws)
    type_key = next((k for k in hm if "Infrastructure" in k or "Type" in k), None)
    elem_key = next((k for k in hm if "Key Elements" in k or "Capabilities" in k), None)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = s(row[hm[type_key]]) if type_key is not None else None
        if not t: continue
        out.append({
            "id": slug(t), "type": t,
            "elements": s(row[hm[elem_key]]) if elem_key is not None and hm[elem_key] < len(row) else None,
        })
    return out

# ─── 8. SPACE STATS (carry-forward stat groups) ─────────────────────────────
def build_stats():
    ws = wb["Space Stats"]; hm = header_map(ws)
    i_stat = hm.get("Stat"); i_unit = hm.get("Unit"); i_val = hm.get("Value"); i_src = hm.get("Sources")
    # label column sits between Unit and Value (e.g. 'Falcon9')
    out, cur_stat = [], None
    for row in ws.iter_rows(min_row=2, values_only=True):
        stat = s(row[i_stat]) if i_stat is not None and i_stat < len(row) else None
        if stat: cur_stat = stat
        unit = s(row[i_unit]) if i_unit is not None and i_unit < len(row) else None
        val  = s(row[i_val])  if i_val  is not None and i_val  < len(row) else None
        src  = s(row[i_src])  if i_src  is not None and i_src  < len(row) else None
        label = None
        if i_unit is not None and i_val is not None and i_val - i_unit > 1:
            label = s(row[i_unit + 1]) if i_unit + 1 < len(row) else None
        if val is None and unit is None: continue
        out.append({"stat": cur_stat, "unit": unit, "label": label, "value": val, "source": src})
    return out

# ─── 9. LAUNCH LOG ──────────────────────────────────────────────────────────
def build_launch_log():
    ws = wb["Launch"]; hm = header_map(ws)
    cols = [("year","Date"),("missionDate","Mission"),("name","Name"),
            ("altNames","Alternate Name(s)"),("type","Type"),
            ("applications","Main Application(s)"),("instruments","Mission Instrument(s)"),
            ("gsd","Main mission min GSD"),("launchDate","Launch Date"),
            ("launcher","Baseline Launcher")]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = s(get(row, hm, "Name"))
        if not name: continue
        obj = {}
        for key, hk in cols:
            v = get(row, hm, hk)
            obj[key] = str(v) if v is not None and not is_err(v) else None
        out.append(obj)
    return out

# ─── 10. NARRATIVES (long-form text sheets) ─────────────────────────────────
def build_narratives():
    out = {}
    for sheet, key in [("Launch Facilities","launchFacilities"),
                       ("Other Facilities","testCentres"),
                       ("EU launchers","euProviders")]:
        ws = wb[sheet]
        paras = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            for c in row:
                t = s(c)
                if t and len(t) > 2:
                    paras.append(t)
        out[key] = paras
    return out

# ─── Run ────────────────────────────────────────────────────────────────────
print("ETL Database.xlsx -> src/data/*.json")
save("launchers.json", build_launchers())
save("launcherClass.json", build_class())
save("csgLaunchpads.json", build_launchpads())
save("spaceports.json", build_spaceports())
save("propulsion.json", build_propulsion())
save("criticalMaterials.json", build_materials())
save("groundInfra.json", build_ground_infra())
save("spaceStats.json", build_stats())
save("launchLog.json", build_launch_log())
save("narratives.json", build_narratives())

# Quick QA summary
launchers = json.load(open(os.path.join(OUT,"launchers.json"), encoding="utf-8"))
eu = [l for l in launchers if l.get("isEU")]
pads = json.load(open(os.path.join(OUT,"csgLaunchpads.json"), encoding="utf-8"))
print("\nQA:")
print(f"  launchers total={len(launchers)}  EU={len(eu)}")
print(f"  pads with coords={sum(1 for p in pads if p.get('coords'))}/{len(pads)}")
print("  sample EU launchers:", ", ".join(l['name'] for l in eu[:12]))
print("Done.")
